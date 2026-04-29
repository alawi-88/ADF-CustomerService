"""Build a 30k-row synthetic xlsx by amplifying the real 3,190-row export."""
import random
from datetime import timedelta
from openpyxl import load_workbook, Workbook

src_path = "/sessions/vigilant-elegant-hawking/mnt/uploads/بيانات مشاركة المستخدمين.xlsx"
out_path = "/tmp/adf/scale_30k.xlsx"
TARGET = 30000

wb = load_workbook(src_path, data_only=True)
ws = wb.active
rows = list(ws.iter_rows(values_only=True))
header = rows[0]
data = rows[1:]
print(f"Source: {len(data)} rows")

new_wb = Workbook()
new_ws = new_wb.active
new_ws.append(list(header))

# Generate TARGET rows by cycling source data, perturbing IDs and dates
random.seed(42)
max_id = max(r[0] for r in data if r[0] is not None)
copies = TARGET // len(data) + 1
written = 0
for c in range(copies):
    for r in data:
        if written >= TARGET: break
        new_id = (r[0] or 0) + (c * 10_000_000) + 1
        new_date = r[4]
        if new_date and c > 0:
            try:
                new_date = new_date - timedelta(days=random.randint(0, 365))
            except Exception:
                pass
        new_ws.append([new_id, r[1], r[2], r[3], new_date])
        written += 1
    if written >= TARGET: break

new_wb.save(out_path)
print(f"Wrote {written} rows to {out_path}")
